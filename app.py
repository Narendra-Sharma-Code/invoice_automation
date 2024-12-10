from flask import Flask, send_file, render_template, request, jsonify  # Add jsonify here
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import load_workbook
from fuzzywuzzy import fuzz, process
import re
import os
import pandas as pd 
import io
app = Flask(__name__)

# Ensure the output directory exists
OUTPUT_DIR = "outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Styles for the Excel file
BOLD_FONT = Font(bold=True, size=12)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
DARK_BOTTOM_BORDER = Border(bottom=Side(style='medium'))
DARK_VERTICAL_BORDER = Border(right=Side(style='medium'))
DARK_BOTTOM_EXTENDED_BORDER = Border(bottom=Side(style='medium'), right=Side(style='medium'))
VERTICAL_BOLD_BORDER = Border(left=Side(style='medium'))  # For bold vertical borders

# Helper function to set cell value, style, and borders
def set_cell(ws, cell, value, font=None, alignment=None, border=None):
    ws[cell] = value
    if font:
        ws[cell].font = font
    if alignment:
        ws[cell].alignment = alignment
    if border:
        ws[cell].border = border

@app.route('/')
def upload_file():
    return render_template('upload.html')  # HTML template for file upload

@app.route('/process', methods=['POST'])
def process_file():
    print(f"Request method: {request.method}")

   # Check if an input file and form data were uploaded
    if 'file' not in request.files or not request.form:
        return "No file or form data uploaded", 400

    input_file = request.files['file']
    if input_file.filename == '':
        return "No selected file", 400

    # Load the input Excel file
    wb_input = load_workbook(input_file)
    ws_input = wb_input.active  # Assuming data is in the first sheet


    # Create a new Excel workbook and active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Setting up the heading
    ws.merge_cells('A1:D1')
    set_cell(ws, 'A1', "INVOICE", font=BOLD_FONT, alignment=CENTER_ALIGN)

 # Get user inputs from the form
    invoice_no = request.form.get('invoice_no', 'Default Invoice')
    exporter_name = request.form.get('exporter_name', 'Default Exporter')

 # Populate dynamic fields
    set_cell(ws, 'E3', f" {invoice_no}", font=BOLD_FONT, alignment=LEFT_ALIGN)
    set_cell(ws, 'H3', f"{exporter_name}", font=BOLD_FONT, alignment=LEFT_ALIGN)
 # Adding a horizontal bold bottom border between Row 5 and Row 6 from Column D to P
    for col in range(4, 16):  # Columns D (4) to P (16)
        cell = ws.cell(row=6, column=col)
        cell.border = Border(bottom=Side(style='medium'))  # Apply bold bottom border

    # Create a horizontal line between row 31 and 32 from column A to P
    for col in range(1, 16):  # Columns A (1) to P (16)
        cell = ws.cell(row=31, column=col)
        cell.border = DARK_BOTTOM_BORDER

     # Create a horizontal line between row 62 and 63 from column A to P
    for col in range(1, 16):  # Columns A (1) to P (17)
        cell = ws.cell(row=62, column=col)

  # Create a horizontal line between row 63  from column A to P
    for col in range(1, 17):  # Columns A (1) to P (17)
        cell = ws.cell(row=63, column=col)

  # Create a horizontal line between row 83  from column A to P
    for col in range(1, 17):  # Columns A (1) to P (17)
        cell = ws.cell(row=83, column=col)


        cell.border = DARK_BOTTOM_BORDER
  # Create a horizontal line between row 62 and 63 from column A to P
    for col in range(1, 16):  # Columns A (1) to P (17)
        cell = ws.cell(row=61, column=col)
        cell.border = DARK_BOTTOM_BORDER

 # Create a horizontal line between row 62 and 63 from column A to P
    for col in range(3, 16):  # Columns A (1) to P (17)
        cell = ws.cell(row=77, column=col)
        cell.border = DARK_BOTTOM_BORDER
    # Add "Banker :" in row 26, column 5 in bold letters
    #set_cell(ws, 'E26', "Banker :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Create a horizontal line between row 26 from column D to P
    for col in range(2, 5):  # Columns B (2) to E (5)
        cell = ws.cell(row=26, column=col)
        cell.border = DARK_BOTTOM_BORDER

  
   

    set_cell(ws, 'E16', "Buyer(if other than consignee)", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding a bold bottom border between row 14 and 15, from column D to P
    for col in range(4, 17):  # Columns D (4) to P (16)
        cell = ws.cell(row=14, column=col)
        cell.border = DARK_BOTTOM_BORDER

    ws.merge_cells('E7:G7')
    set_cell(ws, 'E7', "Ref No:-IN/23/JBW/177", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Merge columns 5, 6, and 7 in row 10
    ws.merge_cells('E8:G8')
    set_cell(ws, 'E8', "Buyer's Ord No. & Date Ref.", font=BOLD_FONT, alignment=LEFT_ALIGN)
    
    # Adding a bold bottom border between row 9 and 10, from column E to P
    for col in range(5, 17):  # Columns D (4) to P (16)
        cell = ws.cell(row=9, column=col)
        cell.border = DARK_BOTTOM_BORDER
   
    ws.merge_cells('E2:F2')
    set_cell(ws, 'E2', "Invoice No. & Date :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding "Other Reference(s)" in Columns E and F, Row 10 in bold letters
    ws.merge_cells('E10:F10')
    set_cell(ws, 'E10', "Other Reference(s)", font=BOLD_FONT, alignment=LEFT_ALIGN)


    ws.merge_cells('H2:I2')
    set_cell(ws, 'H2', "Exporter's Ref :", font=BOLD_FONT, alignment=LEFT_ALIGN)

   

   


    # Adding a vertical bold line between Columns D and E for Rows 2 to 31
    for row in range(78, 84):  # Rows 2 to 31
        ws.cell(row=row, column=2).border = Border(right=Side(style='medium'))  # Bold right border for Column D
 
    # Adding a vertical bold line between Columns D and E for Rows 2 to 31
    for row in range(2, 32):  # Rows 2 to 31
        ws.cell(row=row, column=4).border = Border(right=Side(style='medium'))  #

    # Adding a vertical bold line between Columns P and Q for Rows 2 to 31
    for row in range(1, 84):  # Rows 2 to 31
        ws.cell(row=row, column=16).border = Border(right=Side(style='medium'))  # Bold right border for Column D

    # Exporter Section (Heading Only)
    ws.merge_cells('A2:D2')
    set_cell(ws, 'A2', "Exporter :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Factory Address Section
    ws.merge_cells('A3:B7')
    set_cell(ws, 'A3', "Factory Address :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    ws.merge_cells('C3:D7')
    set_cell(ws, 'C3', ("UNIDESIGN ELITE JEWELLERY PVT LTD.\n"				
                        "Survay No.280091,Mahendra Brothers Exports Pvt Ltd\n"					
	                    "Gandevi Road,Jamalpore\n"				
	                    "At Village Navasari,At Taluka Navasari,At Distric Navasari\n"					
	                    "PIN NO.396445\n"					
	                    "GST-24AAACK3499E1ZL"), alignment=LEFT_ALIGN)

    # Sales Office Address Section
    ws.merge_cells('A9:B9')
    set_cell(ws, 'A9', "Sales Office Address :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    ws.merge_cells('C9:D14')
    set_cell(ws, 'C9', ("UNIDESIGN ELITE JEWELLERY PVT LTD.\n"			
                        "Plot No. D-7/1, 1st Floor, Asian House.\n"				
                        "RD NO. 16, Opp. Prasad LAB,\n"  				
	                    "MIDC Ahdheri (E) Mumbai - 400093\n" 				
                        "GST-27AAACK3499E2ZE\n"				
                        "LUT ARN No-AD2709230019051 Date:- 01/4/2023 To 31/03/2024"), alignment=LEFT_ALIGN)

    # Adding a dark bottom border for row 14
    for col in range(1, 5):  # Columns A to D
        cell = ws.cell(row=14, column=col)
        cell.border = DARK_BOTTOM_BORDER

    # Consignee Section
    ws.merge_cells('A16:B16')
    set_cell(ws, 'A16', "Consignee :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    ws.merge_cells('C16:D22')
    set_cell(ws, 'C16', ("UNIDESIGN JEWELLERY PVT.LTD. UNIT III\n"
                         "PLOT # 4, 5 & 6 (Part)"
                         "SEEPZ, SEZ, Andheri (East).\n"
                         "Mumbai-400096.INDIA\n"
                         "GST : 27AAACU0572G1ZH\n"
                         "Pan No. : AAACU0572G"), alignment=LEFT_ALIGN)

    # Adding a dark bottom border for row 22
    for col in range(1, 5):  # Columns A to D
        cell = ws.cell(row=22, column=col)
        cell.border = DARK_BOTTOM_BORDER

    # Pre-Carriage Section
    ws.merge_cells('A24:B24')
    set_cell(ws, 'A24', "Pre-Carriage by :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding "Place of Receipt by Pre-Carrier :" in column 3, row 24
    set_cell(ws, 'C24', "Place of Receipt by Pre-Carrier :", font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adding "N.A" in column A just below "Pre-Carriage by :"
    set_cell(ws, 'A25', "N.A", alignment=LEFT_ALIGN)

    # Adding "Final Destination :" in column 3, row 27
    set_cell(ws, 'C27', "Final Destination :", font=BOLD_FONT, alignment=LEFT_ALIGN)


    # Adding extended bold bottom border for rows 25 and 26
    for col in range(1, 3):  # Columns A to C
        cell = ws.cell(row=25, column=col)
        cell.border = DARK_BOTTOM_EXTENDED_BORDER


    # Adding extended bold bottom border for rows 26 and 27
    for col in range(1, 3):  # Columns A to C
        cell = ws.cell(row=26, column=col)
        cell.border = DARK_BOTTOM_EXTENDED_BORDER

    # Adding "Port of Discharge :" in row 27
    set_cell(ws, 'A27', "Port of Discharge :", font=BOLD_FONT, alignment=LEFT_ALIGN)
    set_cell(ws, 'A28', "N.A.", alignment=LEFT_ALIGN)
   

    # Adding bold vertical border between columns B and D for rows 22 to 30
    for row in range(22, 32):  # Rows 22 to 30
        ws.cell(row=row, column=3).border = VERTICAL_BOLD_BORDER  # Bold vertical line between columns B

    # Add a horizontal line between row 22 and 23 from columns D to P
    for col in range(4, 16):  # Columns D to P (4 to 16)
        cell = ws.cell(row=22, column=col)
        cell.border = Border(bottom=Side(style='medium'))
    
    # Add a horizontal line between row 1 and 2 from columns D to P
    for col in range(4, 17):  # Columns D to P (4 to 16)
        cell = ws.cell(row=1, column=col)
        cell.border = Border(bottom=Side(style='medium'))

    # Adding horizontal line (bottom border) between row 25 and 26 from column B to E
    for col in range(2, 5):  # Columns B to E
        cell = ws.cell(row=25, column=col)
        cell.border = Border(bottom=Side(style='medium'))

   
    
    # Adding "Terms of Delivery & Payment :" in row 23, columns E, F, and G in bold letters
    ws.merge_cells('E23:G23')  # Merge columns E, F, and G in row 23
    set_cell(ws, 'E23', "Terms of Delivery and Payment", font=BOLD_FONT, alignment=LEFT_ALIGN)

# Adding "Payment Term: Immediate" in row 24, columns E, F, and G
    ws.merge_cells('E24:G24')  # Merge columns E, F, and G in row 24
    set_cell(ws, 'E24', " Job-Work-Return: ", font=None, alignment=LEFT_ALIGN)


    ws.merge_cells('E26:k26')
    set_cell(ws, 'E26',
              f"Labour charges for JOB WORK completed by us for our invoice no: {invoice_no} \n"
              f"against your job work invoice no{exporter_name}.\n"  
              "Vide REQUEST ID No.  492400120343 Challan no: 8000473")

    start_row = ws.max_row + 2  # Start below the existing formatted data

    # First column: "Marks & Nos./ Container No."
    ws.merge_cells('A33:A34')
    set_cell(ws,'A33', "Marks & Nos./ Container No.", font=BOLD_FONT, alignment=CENTER_ALIGN)
   
    set_cell(ws, 'B33', "No. & Kind of Pkgs", font=BOLD_FONT, alignment=LEFT_ALIGN)
    
    
    set_cell(ws, 'C33', "Total GMS", font=BOLD_FONT, alignment=LEFT_ALIGN)

    set_cell(ws, 'C62', "0.000", font=BOLD_FONT, alignment=LEFT_ALIGN)

    set_cell(ws, 'B35',"Description Of Goods", font=BOLD_FONT, alignment=CENTER_ALIGN)
    set_cell(ws, 'A35', "Sr NO.", font=BOLD_FONT, alignment=LEFT_ALIGN)
    set_cell(ws, 'A37', "1", font=BOLD_FONT, alignment=LEFT_ALIGN)
    
    ws.merge_cells('B37:B39')
    set_cell(ws, 'B37', 
                f"Labour charges for JOB WORK completed by us for our invoice no:{invoice_no} \n"
                f"against your job work invoice no{exporter_name}\n"  
                "Vide REQUEST ID No.  492400120343 Challan no: 8000473", font=BOLD_FONT, alignment=LEFT_ALIGN)

    ws.merge_cells('B60:B61')
    set_cell(ws,'B60',
             "EQUI Amount -Rs 258839.58\n "
             "Conv Rate         82.20 ",font=BOLD_FONT, alignment=LEFT_ALIGN)
  
    set_cell(ws, 'D33', "Total PCS", font=BOLD_FONT, alignment=LEFT_ALIGN)
    set_cell(ws, 'D62', "0", font=BOLD_FONT, alignment=LEFT_ALIGN)

    set_cell(ws, 'E33', "Value IN US$", font=BOLD_FONT, alignment=LEFT_ALIGN)
    set_cell(ws, 'E62', " $-   ", font=BOLD_FONT, alignment=LEFT_ALIGN)


    ws.merge_cells('A64:D64')
    set_cell(ws, 'A64', ("Amount Chargable -(in Words)-US$ Three Thousand One Hundred Forty Eight and Cent Ninety Only"),font=BOLD_FONT, alignment=LEFT_ALIGN)

    set_cell(ws,'A66',"I.E.Code No.",font=BOLD_FONT, alignment=LEFT_ALIGN )

    set_cell(ws,'C78', "Undesign Elite Jewellery PVT LTD",font=BOLD_FONT, alignment=LEFT_ALIGN)
    set_cell(ws,'C81', "Authorised Signatory",font=BOLD_FONT, alignment=LEFT_ALIGN)


    set_cell(ws,'A82', "Declaration: ",font=BOLD_FONT, alignment=LEFT_ALIGN)

    ws.merge_cells('A83:B83')
    set_cell(ws,'A83', 
             "We Decalre that this Platform Invoice shows the actual price of the\n"
             "Goods described and that all particulars are true and correct."
             ,font=BOLD_FONT, alignment=LEFT_ALIGN)

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
   
    df = pd.read_excel(input_file, sheet_name='ppl-Packing List', header=3)

    # Strip leading/trailing spaces from column names
    df.columns = df.columns.str.strip()


    if 'Labour' in df.columns:
        # Ensure the 'Labour' column contains numeric data
        # Non-numeric entries are converted to NaN
        df['Labour'] = pd.to_numeric(df['Labour'], errors='coerce')

        # Calculate the sum of all numeric values in the 'Labour' column
        labour_sum = df['Labour'].sum(skipna=True)
        # print( df['Labour'].to_string())

        # Output the final sum
        print(f"Final Sum of Labour Column: {labour_sum}")
    else:
        # Handle the case where the 'Labour' column is missing
        labour_sum = 0
        print("No 'Labour' column found. Setting labour_sum to 0.")

    # Save the workbook to a file
    output_file = os.path.join(OUTPUT_DIR, "Formatted_Invoice.xlsx")
    wb.save(output_file)
    # Open the output file and update cell E62
    wb = load_workbook(output_file)
    ws = wb.active
    ws['E62'] = labour_sum  # Place the sum in cell E62
    wb.save(output_file)
    print(f"Total placed at E62 and file saved as '{output_file}'.")
    print(f"Expected Sum: {3148.9}, Calculated Sum: {labour_sum}")


    # Save the workbook to a file
    output_file = os.path.join(OUTPUT_DIR, "Formatted_Invoice.xlsx")
    wb.save(output_file)

    # Return the file as a download
    return send_file(output_file, as_attachment=True, download_name="Formatted_Invoice.xlsx")


if __name__ == '__main__':
    app.run(debug=True)
